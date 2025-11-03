import openpyxl
from pathlib import Path
from datetime import datetime
import pandas as pd

# =========================
# ユーティリティ
# =========================
def get_unique_path(path: Path) -> Path:
    """同名ファイルが存在する場合、(1),(2)... を付けて重複回避する"""
    if not path.exists():
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    i = 1
    while True:
        candidate = parent / f"{stem}({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1

def parse_row_list_file(filepath: Path):
    """row_list_*.txt を解析し、(diff_URL, filterling_URL) のリストを返す"""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    blocks = [b.strip() for b in content.split("--- row_start ---") if b.strip()]
    results = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        diff_urls, filter_urls = [], []
        if "diff_URL:" in lines:
            diff_start = lines.index("diff_URL:") + 1
            if "filterling_URL:" in lines:
                filter_start = lines.index("filterling_URL:")
                diff_urls = lines[diff_start:filter_start]
                filter_urls = lines[filter_start + 1:]
            else:
                diff_urls = lines[diff_start:]
                filter_urls = []
        # ノイズ除去
        filter_urls = [u for u in filter_urls if u != "（なし）" and not u.startswith("--- row_start ---")]
        results.append(("\n".join(diff_urls), "\n".join(filter_urls)))
    return results

def find_latest_sparse_log(base_dir: Path, genre: str) -> Path | None:
    """
    base_dir（=ファイルAがあるディレクトリ）直下 log_Searched/ から、
    searched(genre)_Keyword-list_*__log_*.xlsx（or .csv） の最新を返す
    """
    log_dir = base_dir / "log_Searched"
    if not log_dir.exists():
        return None
    patterns = [
        f"searched({genre})_Keyword-list_*__log_*.xlsx",
        f"searched({genre})_Keyword-list_*__log_*.csv",
    ]
    cands = []
    for pat in patterns:
        cands.extend(sorted(log_dir.glob(pat)))
    if not cands:
        return None
    def parse_ts(p: Path):
        try:
            ts = p.stem.split("__log_")[-1]
            return datetime.strptime(ts, "%Y%m%d-%H%M%S")
        except Exception:
            return datetime.fromtimestamp(p.stat().st_mtime)
    cands.sort(key=parse_ts, reverse=True)
    return cands[0]

# =========================
# ① 同階層の「フォルダ」を列挙して選択（allなし・複数番号OK）
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent
dirs_1depth = sorted([p for p in SCRIPT_DIR.iterdir() if p.is_dir()])
if not dirs_1depth:
    raise FileNotFoundError("同じフォルダ直下にサブフォルダが見つかりません。")

print("転記対象のフォルダを選択してください:")
for i, d in enumerate(dirs_1depth, start=1):
    print(f"{i}: {d.name}")
n_dirs = len(dirs_1depth)
raw_dir_pick = input(f"番号（1〜{n_dirs}。カンマ区切りで複数可）: ").strip()
dir_idxs = sorted({int(x.strip()) for x in raw_dir_pick.split(",") if x.strip().isdigit()})
if not dir_idxs:
    raise ValueError(f"フォルダ番号の入力が不正です。1〜{n_dirs} の範囲で指定してください。")
target_dirs = [dirs_1depth[i-1] for i in dir_idxs if 1 <= i <= n_dirs]

# =========================
# ② 各フォルダで row_list_*.txt を選択（allなし・複数番号OK）
# ③ 転記先 Excel は 接頭辞 "Keyword-list_" を自動選択
# ④ TXTの内容を対象シートに転記し、対象シートだけの新規ブックを
#    「trsc(〇〇)_」形式で保存。
#    転記先行は、row_list_* の * 部分（=小分類名）に合致する
#    最新のスパースログ D の processed_at 行を**参照**して決定。
# =========================
for base_dir in target_dirs:
    # --- row_list_*.txt を列挙 ---
    row_list_files = sorted(base_dir.glob("row_list_*.txt"))
    if not row_list_files:
        print(f"[WARN] フォルダ '{base_dir.name}' に row_list_*.txt が見つかりません。スキップします。")
        continue

    print("\n" + "="*72)
    print(f"▶ フォルダ: {base_dir.name}")
    print("転記元の TXT ファイルを選択してください:")
    for i, f in enumerate(row_list_files, start=1):
        print(f"{i}: {f.name}")
    n_txt = len(row_list_files)
    raw_txt_pick = input(f"番号（1〜{n_txt}。カンマ区切りで複数可）: ").strip()
    txt_idxs = sorted({int(x.strip()) for x in raw_txt_pick.split(",") if x.strip().isdigit()})
    if not txt_idxs:
        print(f"[WARN] TXTの番号入力が空 or 不正です（1〜{n_txt}）。スキップします。")
        continue
    target_txts = [row_list_files[i-1] for i in txt_idxs if 1 <= i <= n_txt]

    # --- 転記先の Excel を自動選択: 接頭辞 "Keyword-list_" のみ対象 ---
    excel_candidates = sorted(base_dir.glob("Keyword-list_*.xlsx"))
    if not excel_candidates:
        print(f"[WARN] フォルダ '{base_dir.name}' に 'Keyword-list_*.xlsx' が見つかりません。スキップします。")
        continue

    print("\n自動選択された転記先 Excel（Keyword-list_*）:")
    for i, f in enumerate(excel_candidates, start=1):
        print(f"{i}: {f.name}")

    # --- TXT群 × Excel群 で処理 ---
    for excel_path in excel_candidates:
        print("\n" + "-"*72)
        print(f"▶ 転記先 Excel: {excel_path.name}")

        for txt_path in target_txts:
            sheet_name_candidate = txt_path.stem.replace("row_list_", "").strip()
            print(f"  - TXT: {txt_path.name} → シート候補: '{sheet_name_candidate}'")

            # 1) TXT読み込み
            row_sets = parse_row_list_file(txt_path)

            # 2) Excel を読み込み
            wb = openpyxl.load_workbook(excel_path)

            # 3) シート名 完全一致（前後空白トリム）
            match_name = next((n for n in wb.sheetnames if n.strip() == sheet_name_candidate), None)
            if not match_name:
                print(f"    ✖ シート '{sheet_name_candidate}' が見つかりません。スキップ。")
                continue

            ws = wb[match_name]
            print(f"    ✅ 対象シート: {ws.title}")

            # 4) 必須列確認（無ければ作成）
            headers = {cell.value: col_idx for col_idx, cell in enumerate(ws[1], start=1)}
            changed = False
            for need in ["diff_URL", "filterling_URL"]:
                if need not in headers:
                    ws.cell(row=1, column=ws.max_column + 1).value = need
                    headers[need] = ws.max_column
                    changed = True
            if changed:
                headers = {cell.value: col_idx for col_idx, cell in enumerate(ws[1], start=1)}

            diff_col = headers["diff_URL"]
            filter_col = headers["filterling_URL"]

            # 5) 参照ログから「転記先行」を決定
            genre = sheet_name_candidate
            latest_log = find_latest_sparse_log(excel_path.parent, genre)
            if latest_log is None:
                print("    ⚠ 参照ログが見つからないため、従来どおり先頭から順に転記します。")
                target_rows = list(range(2, 2 + len(row_sets)))  # 2行目から
            else:
                print(f"    ↪ 参照ログ: {latest_log.name}")
                if latest_log.suffix.lower() == ".xlsx":
                    dfl = pd.read_excel(latest_log, sheet_name=sheet_name_candidate)
                else:
                    dfl = pd.read_csv(latest_log)
                if "processed_at" not in dfl.columns:
                    print("    ⚠ ログに 'processed_at' 列がないため、先頭から順に転記します。")
                    target_rows = list(range(2, 2 + len(row_sets)))
                else:
                    processed_idx = [int(i) for i, v in enumerate(dfl["processed_at"].fillna("").tolist()) if str(v).strip() != ""]
                    target_rows = [i + 2 for i in processed_idx]  # 1行目がヘッダ
                    if not target_rows:
                        print("    ⚠ ログに処理行が見つからないため、先頭から順に転記します。")
                        target_rows = list(range(2, 2 + len(row_sets)))

            # 転記数は行リストとTXT側の最小に合わせる
            n_write = min(len(row_sets), len(target_rows))
            if n_write == 0:
                print("    ⚠ 転記対象がありません。スキップ。")
                continue

            # 6) 指定行に転記（上書き）
            written = 0
            for k in range(n_write):
                tgt_row = target_rows[k]
                diff_str, filter_str = row_sets[k]
                ws.cell(row=tgt_row, column=diff_col).value = diff_str
                ws.cell(row=tgt_row, column=filter_col).value = filter_str
                written += 1

            print(f"    📝 書き込んだ行数: {written}（TXT {len(row_sets)}件 / ログ行 {len(target_rows)}件 → 使用 {n_write}件）")

            # 7) 対象シートだけの新規ブックを作成し保存（trsc(〇〇)_）
            new_wb = openpyxl.Workbook()
            default_ws = new_wb.active
            new_wb.remove(default_ws)

            nws = new_wb.create_sheet(ws.title)
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for c_idx, v in enumerate(row, start=1):
                    nws.cell(row=r_idx, column=c_idx).value = v

            out_name = f"trsc({ws.title})_{excel_path.name}"
            out_path = get_unique_path(excel_path.parent / out_name)
            new_wb.save(out_path)
            print(f"    💾 保存（対象シートのみ）: {out_path}")

    print("\n完了しました。")
